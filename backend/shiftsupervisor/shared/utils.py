from openpyxl.styles import Font, Color

font_style_title_table = Font(size=14, color="FF0000")


def write_unit_data_in_xlsx(sheet, start_row, column, unit_data, fields):
    for i, field_name in enumerate(fields, start=start_row):
        value = getattr(unit_data, field_name, "")
        sheet.cell(row=i, column=column, value=value)


def write_unit_name_in_xlsx(sheet, title_table, start_row, column, unit_data, fields):
    unit_cell = sheet.cell(row=start_row, column=1, value=title_table)
    unit_cell.font = font_style_title_table
    for i, field_name in enumerate(fields, start=start_row + 1):
        sheet.cell(row=i, column=column, value=field_name)


def creat_title_sheet(sheet, start_row, column=()):
    write_unit_name_in_xlsx(
        sheet, start_row=4, column=1, unit_data=data.u500_data, fields=u500_fields
    )
    u600_cell = sheet.cell(row=4 + len(u500_fields), column=1, value="U600")
    u600_cell.font = font_style_title_table


def get_unit_filds_for_xlsx(unit, exclude_columns):
    return [
        field.name for field in unit._meta.fields if field.name not in exclude_columns
    ]
