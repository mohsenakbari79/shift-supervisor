import jdatetime
from openpyxl import Workbook
from django.http import StreamingHttpResponse
from px.models import DailyDataPX
from datetime import datetime
from io import BytesIO
from rest_framework.decorators import api_view
from openpyxl.styles import Font, Color

from shared.utils import (
    write_unit_data_in_xlsx,
    write_unit_name_in_xlsx,
    creat_title_sheet,
    get_unit_filds_for_xlsx,
)


@api_view(["GET"])
def export_to_excel(request):

    # دریافت داده‌ها با استفاده از select_related برای کاهش تعداد کوئری‌ها
    daily_data = DailyDataPX.objects.select_related(
        "user", "u400_data", "u700_data", "u800_data", "u950_data", "u970_data"
    ).order_by("date")

    workbook = Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    current_month = None
    column_index = 2

    for data in daily_data:
        shamsi_date = jdatetime.date.fromgregorian(date=data.date)
        month_name = shamsi_date.strftime("%Y-%m")

        if current_month != month_name:
            # اگر ماه جدید باشد، یک شیت جدید ایجاد می‌شود
            current_month = month_name
            sheet = workbook.create_sheet(title=month_name)
            column_index = 2

            sheet["A1"] = "UNIT WISE STATUS:"
            sheet["A2"] = "FEED/PRODUCT CAP."

            # دریافت فیلدهای مورد نیاز برای هر واحد
            u400_fields = get_unit_filds_for_xlsx(data.u400_data, ("id", "date"))
            u700_fields = get_unit_filds_for_xlsx(data.u700_data, ("id", "date"))
            u800_fields = get_unit_filds_for_xlsx(data.u800_data, ("id", "date"))
            u950_fields = get_unit_filds_for_xlsx(data.u950_data, ("id", "date"))
            u970_fields = get_unit_filds_for_xlsx(data.u970_data, ("id", "date"))

            # نوشتن نام واحدها و داده‌ها در شیت
            row_index = 3
            row_index = write_unit_name_in_xlsx(sheet, "U400", row_index, 1, data.u400_data, u400_fields)
            row_index = write_unit_name_in_xlsx(sheet, "U700", row_index, 1, data.u700_data, u700_fields)
            row_index = write_unit_name_in_xlsx(sheet, "U800", row_index, 1, data.u800_data, u800_fields)
            row_index = write_unit_name_in_xlsx(sheet, "U950", row_index, 1, data.u950_data, u950_fields)
            row_index = write_unit_name_in_xlsx(sheet, "U970", row_index, 1, data.u970_data, u970_fields)

        # افزودن نام کاربر و تاریخ در ستون‌های مختلف
        sheet.cell(row=1, column=column_index, value=getattr(data.user, 'first_name', 'نام‌وجود‌ندارد'))
        sheet.cell(row=2, column=column_index, value=str(shamsi_date))

        # نوشتن داده‌های هر واحد با تنظیم اندیس سطر به صورت داینامیک
        row_index = 4
        row_index = write_unit_data_in_xlsx(sheet, row_index, column_index, data.u400_data, u400_fields)
        row_index = write_unit_data_in_xlsx(sheet, row_index, column_index, data.u700_data, u700_fields)
        row_index = write_unit_data_in_xlsx(sheet, row_index, column_index, data.u800_data, u800_fields)
        row_index = write_unit_data_in_xlsx(sheet, row_index, column_index, data.u950_data, u950_fields)
        row_index = write_unit_data_in_xlsx(sheet, row_index, column_index, data.u970_data, u970_fields)

        column_index += 1

    # تنظیم عرض ستون‌ها برای تمام شیت‌ها
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

    # ذخیره کردن فایل اکسل در یک فایل موقتی در حافظه
    with BytesIO() as output:
        workbook.save(output)
        output.seek(0)

        # ساخت پاسخ HTTP با استریم فایل
        response = StreamingHttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="monthly_data_{timestamp}.xlsx"'

    return response
