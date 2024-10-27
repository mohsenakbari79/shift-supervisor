import jdatetime
from openpyxl import Workbook
from django.http import StreamingHttpResponse
from btx.models import DailyDataBtx
from datetime import datetime
import tempfile
import os
from io import BytesIO  # Import BytesIO for in-memory file handling
from rest_framework.decorators import api_view
from openpyxl.styles import Font, Color
from shared.utils import (
    write_unit_data_in_xlsx,
    write_unit_name_in_xlsx,
    creat_title_sheet,
    get_unit_filds_for_xlsx,
)


font_style_title_table = Font(size=14, color="FF0000")


@api_view(["GET"])
def export_to_excel(request):
    daily_data = DailyDataBtx.objects.select_related(
        "user", "u500_data", "u600_data", "u650_data"
    ).order_by("date")

    workbook = Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    current_month = None
    sheet = None
    column_index = 2

    for data in daily_data:
        shamsi_date = jdatetime.date.fromgregorian(date=data.date)
        month_name = shamsi_date.strftime("%Y-%m")

        if current_month != month_name:
            current_month = month_name
            sheet = workbook.create_sheet(title=month_name)
            column_index = 2

            sheet["A1"] = "UNIT WISE STATUS:"
            sheet["A2"] = "FEED/PRODUCT CAP."

            u500_fields = get_unit_filds_for_xlsx(data.u500_data, ("id", "date"))
            u600_fields = get_unit_filds_for_xlsx(data.u600_data, ("id", "date"))
            u650_fields = get_unit_filds_for_xlsx(data.u650_data, ("id", "date"))

            write_unit_name_in_xlsx(
                sheet,
                title_table="U500",
                start_row=3,
                column=1,
                unit_data=data.u500_data,
                fields=u500_fields,
            )
            write_unit_name_in_xlsx(
                sheet,
                title_table="U600",
                start_row=4 + len(u500_fields),
                column=1,
                unit_data=data.u600_data,
                fields=u600_fields,
            )
            write_unit_name_in_xlsx(
                sheet,
                title_table="U650",
                start_row=5 + len(u500_fields) + len(u600_fields),
                column=1,
                unit_data=data.u650_data,
                fields=u650_fields,
            )

        sheet.cell(row=1, column=column_index, value=data.user.first_name)
        sheet.cell(row=2, column=column_index, value=str(shamsi_date))

        row_index = 4
        write_unit_data_in_xlsx(
            sheet, row_index, column_index, data.u500_data, u500_fields
        )
        row_index += len(u500_fields) + 1

        write_unit_data_in_xlsx(
            sheet, row_index, column_index, data.u600_data, u600_fields
        )
        row_index += len(u600_fields) + 1

        write_unit_data_in_xlsx(
            sheet, row_index, column_index, data.u650_data, u650_fields
        )

        column_index += 1

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # حرف ستون
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # 2 را به طول اضافه می‌کند
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Saving the Excel file to a BytesIO object instead of a temporary file
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Move to the beginning of the BytesIO buffer

    # Creating an HTTP response with streaming
    response = StreamingHttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = (
        f'attachment; filename="monthly_data_{timestamp}.xlsx"'
    )

    return response
