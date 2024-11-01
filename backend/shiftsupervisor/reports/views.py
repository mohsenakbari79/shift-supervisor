import jdatetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from django.http import StreamingHttpResponse, HttpResponse
from px.models import DailyDataPX
from btx.models import DailyDataBtx
from reforming.models import DailyDataReforming
from datetime import datetime
from io import BytesIO
from rest_framework.decorators import api_view

from shared.utils import (
    write_unit_data_in_xlsx,
    write_unit_name_in_xlsx,
    get_unit_filds_for_xlsx,
)

@api_view(["GET"])
def export_all_units_to_excel(request):
    try:
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # تنظیم رنگ برای سطرها و ستون‌ها
        header_fill = PatternFill(start_color="FDE0C8", end_color="FDE0C8", fill_type="solid")
        column_fill = PatternFill(start_color="EAF1DD", end_color="EAF1DD", fill_type="solid")
        help_text_fill = PatternFill(start_color="DDDFFF", end_color="DDDFFF", fill_type="solid")  # رنگ متفاوت برای ستون دوم

        # تنظیمات حاشیه‌ها
        thin_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # ایجاد شیء برای چیدمان
        center_alignment = Alignment(horizontal='center', vertical='center')

        # استخراج داده‌ها از هر مدل
        px_data = DailyDataPX.objects.select_related("user", "u400_data", "u700_data", "u800_data", "u950_data", "u970_data").order_by("date")
        u2_data = DailyDataBtx.objects.select_related("user", "u500_data", "u600_data", "u650_data").order_by("date")
        u3_data = DailyDataReforming.objects.select_related("user", "u100_data", "u200_data", "u250_data", "u300_data", "u350_data").order_by("date")

        # ساخت دیکشنری برای گروه‌بندی داده‌ها بر اساس تاریخ
        data_by_date = {}

        def add_data_to_dict(daily_data, unit_prefix):
            for data in daily_data:
                date_key = data.date
                if date_key not in data_by_date:
                    data_by_date[date_key] = {
                        "user": getattr(data.user, "first_name", "نام‌وجود‌ندارد"),
                        "data": {}
                    }

                # اضافه کردن داده‌های هر واحد در هر تاریخ
                unit_data = getattr(data, f"{unit_prefix}_data", None)
                if unit_data:
                    fields = get_unit_filds_for_xlsx(unit_data, ("id", "date"))
                    data_by_date[date_key]["data"][unit_prefix] = (unit_data, fields)

        # پر کردن داده‌ها برای هر مدل
        add_data_to_dict(px_data, "u400")
        add_data_to_dict(px_data, "u700")
        add_data_to_dict(px_data, "u800")
        add_data_to_dict(px_data, "u950")
        add_data_to_dict(px_data, "u970")

        add_data_to_dict(u2_data, "u500")
        add_data_to_dict(u2_data, "u600")
        add_data_to_dict(u2_data, "u650")

        add_data_to_dict(u3_data, "u100")
        add_data_to_dict(u3_data, "u200")
        add_data_to_dict(u3_data, "u250")
        add_data_to_dict(u3_data, "u300")
        add_data_to_dict(u3_data, "u350")

        # گروه‌بندی داده‌ها بر اساس ماه
        data_by_month = {}
        for date, data_dict in data_by_date.items():
            month_key = date.strftime("%Y-%m")
            if month_key not in data_by_month:
                data_by_month[month_key] = []
            data_by_month[month_key].append((date, data_dict))

        # نوشتن داده‌ها برای هر ماه در یک شیت جداگانه
        for month, entries in sorted(data_by_month.items()):
            sheet = workbook.create_sheet(title=f"Report {month}")

            # تنظیم عناوین اولیه
            sheet["A1"] = "UNIT WISE STATUS:"
            sheet["A2"] = "FEED/PRODUCT CAP."

            row_index = 3  # شروع از ردیف سوم
            column_index = 3  # داده‌ها از ستون سوم شروع می‌شوند
            help_text_column_index = 2  # ستون دوم برای help text

            # پر کردن داده‌ها برای هر تاریخ در ماه جاری
            for date, data_dict in entries:
                # نوشتن نام کاربر و تاریخ شمسی در سطر ۱ و ۲
                shamsi_date = jdatetime.date.fromgregorian(date=date)
                sheet.cell(row=1, column=column_index, value=data_dict["user"]).alignment = center_alignment
                sheet.cell(row=2, column=column_index, value=str(shamsi_date)).alignment = center_alignment

                # نوشتن داده‌ها برای هر واحد
                for unit_prefix, (unit_data, fields) in data_dict["data"].items():
                    # نوشتن verbose_name هر فیلد در ستون اول
                    for i, field in enumerate(fields, start=row_index + 1):
                        verbose_name = unit_data._meta.get_field(field).verbose_name or ""
                        name_cell = sheet.cell(row=i, column=1, value=verbose_name)
                        name_cell.fill = column_fill  # رنگ‌آمیزی ستون اول
                        name_cell.border = thin_border  # اعمال حاشیه به سلول‌های ستون اول
                        name_cell.alignment = center_alignment  # وسط‌چین کردن متن

                        help_text = unit_data._meta.get_field(field).help_text
                        help_text_cell = sheet.cell(row=i, column=help_text_column_index, value=help_text or "")
                        help_text_cell.fill = help_text_fill  # رنگ‌آمیزی ستون دوم
                        help_text_cell.border = thin_border  # اعمال حاشیه به سلول‌های help_text
                        help_text_cell.alignment = center_alignment  # وسط‌چین کردن متن

                        value = getattr(unit_data, field, None)
                        data_cell = sheet.cell(row=i, column=column_index, value=value)
                        data_cell.border = thin_border  # اعمال حاشیه به داده‌ها
                        data_cell.alignment = center_alignment  # وسط‌چین کردن متن

                    row_index += len(fields) + 2  # به روزرسانی ردیف برای فاصله

                column_index += 1  # انتقال به ستون بعدی برای تاریخ جدید

                # بازنشانی row_index برای تاریخ جدید
                row_index = 3

            # اعمال رنگ و حاشیه به سطرهای اول و دوم تا آخرین ستون دارای داده
            max_column = sheet.max_column
            for row in [1, 2]:
                for col in range(1, max_column + 1):
                    header_cell = sheet.cell(row=row, column=col)
                    header_cell.fill = header_fill
                    header_cell.border = thin_border
                    header_cell.alignment = center_alignment  # وسط‌چین کردن متن

            # تنظیم رنگ و حاشیه برای ستون اول و دوم تا آخرین سطر دارای داده (حتی اگر خالی باشند)
            max_row = sheet.max_row
            for row in range(3, max_row + 1):
                col1_cell = sheet.cell(row=row, column=1)
                col1_cell.fill = column_fill
                col1_cell.border = thin_border
                col1_cell.alignment = center_alignment  # وسط‌چین کردن متن
                col2_cell = sheet.cell(row=row, column=2)
                col2_cell.fill = help_text_fill
                col2_cell.border = thin_border
                col2_cell.alignment = center_alignment  # وسط‌چین کردن متن

            # تنظیم عرض ستون‌ها برای تمام شیت‌ها
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    cell.border = thin_border  # اعمال حاشیه به همه سلول‌ها
                sheet.column_dimensions[column_letter].width = max_length + 2

        # ذخیره کردن فایل اکسل در یک فایل موقتی در حافظه
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # ساخت پاسخ HTTP با استریم فایل
        response = StreamingHttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="company_combined_report_{timestamp}.xlsx"'

        return response
    except Exception as e:
        print("Error while generating report:", e)
        return HttpResponse(status=500)
