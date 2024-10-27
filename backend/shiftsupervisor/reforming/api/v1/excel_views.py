import jdatetime
from openpyxl import Workbook
from django.http import StreamingHttpResponse
from reforming.models import DailyDataReforming
from datetime import datetime
import tempfile
import os
from io import BytesIO  # Import BytesIO for in-memory file handling
from rest_framework.decorators import api_view
from shared.utils import (
    write_unit_data_in_xlsx,
    write_unit_name_in_xlsx,
    creat_title_sheet,
    get_unit_filds_for_xlsx
)



@api_view(["GET"])
def export_to_excel(request):

    daily_data = DailyDataReforming.objects.select_related(
        "user", "u100_data", "u200_data", "u250_data", "u300_data", "u350_data"
    ).order_by("date")

    workbook = Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    current_month = None
    sheet = None
    column_index = 2

    for data in daily_data:
        shamsi_date = jdatetime.date.fromgregorian(date=data.date)
        month_name = shamsi_date.strftime("%Y-%m")

        if current_month != month_name:
            current_month = month_name
            sheet = workbook.create_sheet(title=month_name)
            column_index = 2

            sheet["A1"] = "UNIT WISE STATUS:"
            sheet["A2"] = "FEED/PRODUCT CAP."

            u100_fields = get_unit_filds_for_xlsx(data.u100_data, ("id", "date"))
            u200_fields = get_unit_filds_for_xlsx(data.u200_data, ("id", "date"))
            u250_fields = get_unit_filds_for_xlsx(data.u250_data, ("id", "date"))
            u300_fields = get_unit_filds_for_xlsx(data.u300_data, ("id", "date"))
            u350_fields = get_unit_filds_for_xlsx(data.u350_data, ("id", "date","capacity_percent")) 

            write_unit_name_in_xlsx(
                sheet,
                title_table="U100",
                start_row=3,
                column=1,
                unit_data=data.u100_data,
                fields=u100_fields,
            )
            write_unit_name_in_xlsx(
                sheet,
                title_table="U200",
                start_row=4 + len(u100_fields),
                column=1,
                unit_data=data.u200_data,
                fields=u200_fields,
            )
            write_unit_name_in_xlsx(
                sheet,
                title_table="U250",
                start_row=5 + len(u100_fields) + len(u200_fields),
                column=1,
                unit_data=data.u250_data,
                fields=u250_fields,
            )
            write_unit_name_in_xlsx(
                sheet,
                title_table="U300",
                start_row=6 + len(u100_fields) + len(u200_fields) + len(u250_fields),
                column=1,
                unit_data=data.u300_data,
                fields=u300_fields,
            )
            write_unit_name_in_xlsx(
                sheet,
                title_table="U350",
                start_row=7
                + len(u100_fields)
                + len(u200_fields)
                + len(u250_fields)
                + len(u300_fields),
                column=1,
                unit_data=data.u350_data,
                fields=u350_fields,
            )

        sheet.cell(row=1, column=column_index, value=data.user.first_name)
        sheet.cell(row=2, column=column_index, value=str(shamsi_date))

        row_index = 4
        write_unit_data_in_xlsx(sheet, row_index, column_index, data.u100_data, u100_fields)

        row_index += len(u100_fields) + 1
        write_unit_data_in_xlsx(sheet, row_index, column_index, data.u200_data, u200_fields)

        row_index += len(u200_fields) + 1
        write_unit_data_in_xlsx(sheet, row_index, column_index, data.u250_data, u250_fields)

        row_index += len(u250_fields) + 1
        write_unit_data_in_xlsx(sheet, row_index, column_index, data.u300_data, u300_fields)

        row_index += len(u300_fields) + 1
        write_unit_data_in_xlsx(sheet, row_index, column_index, data.u350_data, u350_fields)

        column_index += 1

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # حرف ستون
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # 2 را به طول اضافه می‌کند
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Saving the Excel file to a BytesIO object instead of a temporary file
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Move to the beginning of the BytesIO buffer

    # Creating an HTTP response with streaming
    response = StreamingHttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = (
        f'attachment; filename="monthly_data_{timestamp}.xlsx"'
    )

    return response
